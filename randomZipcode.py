import os
import openpyxl
import numpy as np

def randomZipcode():
    directory = os.path.join(os.getcwd(), 'uszips.xlsx')
    wb = openpyxl.load_workbook(directory, read_only=True)
    ws = wb.active
    randomIndex = np.random.randint(2,ws.max_row)
    return ws.cell(row=randomIndex, column=1).value